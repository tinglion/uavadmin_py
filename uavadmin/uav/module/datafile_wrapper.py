import json

import openpyxl
import xlrd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


def load_flight_from_jsonl(fn=r"./data/track_Lshape_0828.txt"):
    data = []
    with open(fn, "r", encoding="utf8") as f:
        for line in f:
            data.append(json.loads(line))
    return data


def load_flight_from_xls(fn=r"./data/radar.xls"):
    position_list = []

    # 打开工作簿
    workbook = xlrd.open_workbook(fn)

    # 获取第一个工作表（索引从0开始）
    sheet = workbook.sheet_by_index(0)

    # 遍历工作表的所有行和列
    for row_idx in range(sheet.nrows):
        if row_idx == 0:
            continue
        position_list.append(
            {
                "topic": sheet.cell(row_idx, 0).value,
                "pos": json.loads(sheet.cell(row_idx, 1).value),
            }
        )

        # row = sheet.row(row_idx)
        # for col_idx in range(sheet.ncols):
        #     cell_value = row[col_idx].value
        #     print(f"Row {row_idx+1}, Column {col_idx+1}: {cell_value}")

    # 如果你知道特定的单元格位置，你也可以直接访问它
    # cell_A1 = sheet.cell(0, 0).value  # 获取第一行第一列的值

    return position_list
