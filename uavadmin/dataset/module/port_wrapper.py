import json
import logging
import os
import re
import shutil
import sys
import traceback
import zipfile
from datetime import datetime

import openpyxl
from django.db import transaction
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from appuav.settings import TEMP_DIR
from uavadmin.dataset.models import DsDataset, DsImage, DsPatient
from uavadmin.dataset.serializers import (
    DsDatasetSerializer,
    DsImageSerializer,
    DsPatientSerializer,
)
from uavadmin.utils.file_utils import get_string_len, zip_folder
from uavadmin.utils.import_export import import_to_data

logger = logging.getLogger(__name__)

default_column_width = 20
max_column_width = 50

dataset_headers = dict(
    header_data=["id", "name", "stage", "description", "status"],
    hidden_header=["id", "name", "stage", "description", "status"],
)

patient_headers = dict(
    header_data=[
        "id",
        "uid",
        "dataset_id",
        "disease_type",
        "crf_result",
        "description",
        "status",
    ],
    hidden_header=[
        "id",
        "uid",
        "dataset_id",
        "disease_type",
        "crf_result",
        "description",
        "status",
    ],
)

image_headers = dict(
    header_data=[
        "id",
        "patient_id",
        "url",
        "url_origin",
        "url_desensitive",
        "scale_factor",
        "ocr_result",
        "ocr_raw",
        "category",
        "preprocess",
        "native_result",
        "native_result_custom",
        "pic_result",
        "description",
        "status",
    ],
    hidden_header=[
        "id",
        "patient_id",
        "url",
        "url_origin",
        "url_desensitive",
        "scale_factor",
        "ocr_result",
        "ocr_raw",
        "category",
        "preprocess",
        "native_result",
        "native_result_custom",
        "pic_result",
        "description",
        "status",
    ],
)


# dataset 有不更新,没有则加
# patient&image，忽略id，直接添加
def import_dataset(zfile):
    flag = True
    msg = "succ"

    file_content = zfile.read()
    # 保存本地文件
    # with open(f"{TEMP_DIR}/{zfile.name}", 'wb') as fp1:
    #     fp1.write(file_content)
    #     fp1.close()

    # 解压缩zip
    with zipfile.ZipFile(zfile, "r") as zip_ref:
        # zip_ref.extractall(f"{TEMP_DIR}")
        # 获取zip文件中的所有文件名
        for file_info in zip_ref.infolist():
            # 解决中文乱码问题：手动指定文件名的编码方式为UTF-8
            file_name = file_info.filename.encode("cp437").decode("gbk")
            if file_name[-1] == "/":
                # 如果是文件夹就直接创建目录
                zip_ref.extract(member=file_info, path=f"{TEMP_DIR}")
            else:
                # 如果是文件，先将文件名从gbk编码转换为utf-8编码
                file_info.filename = file_name
                zip_ref.extract(member=file_info, path=f"{TEMP_DIR}")
    tmp_folder = f"{TEMP_DIR}/{zfile.name.replace('.zip', '').replace('_tmp_ds_', '')}"

    try:
        # 导入dataset
        dataset_list = import_xls_to_data(
            f"{tmp_folder}/dataset.xlsx",
            header_data=dataset_headers["header_data"],
            hidden_header=dataset_headers["hidden_header"],
        )
        if len(dataset_list) < 0:
            return False, "no dataset"
        else:
            with transaction.atomic():
                # 导入dataset,
                dataset = dataset_list[0]
                logger.info(f"{dataset}")

                dataset_id = dataset.get("id", None)
                if dataset_id:
                    dataset_old = DsDataset.objects.filter(id=dataset_id)
                    if len(dataset_old) > 0:
                        return False, "dataset exists"
                dataset_new = DsDataset.objects.create(
                    id=dataset_id,
                    name=dataset.get("name"),
                    stage=dataset.get("stage", 0),
                    description=dataset.get("description"),
                    status=dataset.get("status", 0),
                )

                #    导入patient,
                patient_list = import_xls_to_data(
                    f"{tmp_folder}/patients.xlsx",
                    header_data=patient_headers["header_data"],
                    hidden_header=patient_headers["hidden_header"],
                )
                for patient in patient_list:
                    patient_xls_id = patient.get("uid")
                    patient_new = DsPatient.objects.create(
                        uid=patient.get("uid"),
                        dataset_id=dataset_id,
                        disease_type=patient.get("disease_type"),
                        description=patient.get("description"),
                        status=patient.get("status", 0),
                    )
                    for k in ["crf_result"]:
                        if patient.get(k):
                            setattr(patient_new, k, json.loads(patient.get(k)))
                    patient_new.save()
                    patient_id = patient_new.id

                    #     导入image,
                    image_list = import_xls_to_data(
                        f"{tmp_folder}/patient_{patient_xls_id}.xlsx",
                        header_data=image_headers["header_data"],
                        hidden_header=image_headers["hidden_header"],
                    )
                    for image in image_list:
                        image_url = image.get("url", None)
                        url_origin = image.get("url_origin", None)
                        if not image_url or not url_origin:
                            continue
                        image_new = DsImage.objects.create(
                            patient_id=patient_id,
                            url=image_url,
                            url_origin=url_origin,
                            url_desensitive=image.get("url_desensitive"),
                            scale_factor=image.get("scale_factor", 1),
                            ocr_result=image.get("ocr_result"),
                            category=image.get("category"),
                            description=image.get("description"),
                            status=image.get("status", 0),
                        )
                        for k in [
                            "ocr_raw",
                            "native_result",
                            "native_result_custom",
                            "pic_result",
                        ]:
                            if image.get(k, None):
                                setattr(image_new, k, json.loads(image.get(k)))
                        image_new.save()

                logger.info(f"{dataset_id} {dataset_new.id}")
                #

        # patients = import_xls_to_patient(f"{tmp_folder}/patients.xlsx")

        # for patient in patients:
        #     import_xls_to_image(f"{tmp_folder}/patient_{patient.id}.xlsx")
    except Exception as e:
        logger.error("ERROR")
        traceback.print_exc()
        flag = False
        msg = "数据冲突"

    # 清理文件
    shutil.rmtree(tmp_folder)
    return flag, msg


def import_xls_to_data(fn, header_data, hidden_header):
    # 读取excel 文件
    workbook = openpyxl.load_workbook(fn)
    table = workbook[workbook.sheetnames[0]]
    # 行数
    nrows = table.max_row
    # 列数
    ncols = table.max_column
    # 表头映射
    map_names = {i: table.cell(1, 1 + i).value for i in range(0, ncols)}

    # 创建一个空列表，存储Excel的数据
    tables = []
    for i, row in enumerate(range(nrows)):
        if i == 0:
            continue
        array = {}
        for index in range(0, ncols):
            cell_value = table.cell(1 + row, 1 + index).value
            array[map_names[index]] = cell_value
        # for index, name in enumerate(hidden_header):
        #     cell_value = table.cell(row=row + 1, column=index + 2).value
        #     array[name] = cell_value
        tables.append(array)
    return tables


def export_dataset(id, type="xls"):
    try:
        dataset = DsDataset.objects.get(id=id)
        if not dataset:
            return ErrorResponse(msg="no such dataset")

        # 创建文件夹
        fn_folder = f"{TEMP_DIR}/dataset_{id}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        if not os.path.exists(fn_folder):
            os.makedirs(fn_folder)
        fn_zip = f"{fn_folder}.zip"

        # 导出数据集信息
        export_list_to_xls(
            [DsDatasetSerializer(dataset).data],
            fn=f"{fn_folder}/dataset.xlsx",
            header_data=dataset_headers["header_data"],
            hidden_header=dataset_headers["hidden_header"],
        )
        # with open(f"{fn_folder}/dataset.json", "w", encoding="utf8") as fp1:
        #     json.dump(DsDatasetSerializer(dataset).data, fp1, ensure_ascii=False)
        #     fp1.close()

        # 导出患者列表
        patients = DsPatient.objects.filter(dataset_id=id)
        export_list_to_xls(
            [DsPatientSerializer(patient).data for patient in patients],
            fn=f"{fn_folder}/patients.xlsx",
            header_data=patient_headers["header_data"],
            hidden_header=patient_headers["hidden_header"],
        )

        # # 按患者导出图片列表
        for patient in patients:
            images = DsImage.objects.filter(patient_id=patient.id)
            export_list_to_xls(
                [DsImageSerializer(image).data for image in images],
                fn=f"{fn_folder}/patient_{patient.uid}.xlsx",
                header_data=image_headers["header_data"],
                hidden_header=image_headers["hidden_header"],
            )

        # zip文件夹
        zip_folder(fn_folder, fn_zip)
        shutil.rmtree(fn_folder)

        return fn_zip
    except Exception as e:
        logger.error("export failed")
        traceback.print_exc()
    return None


def export_list_to_xls(
    data: list, fn: str, header_data: list = [], hidden_header: list = []
):
    if not data:
        return False
    default_columns = [attr for attr in dir(data[0]) if not attr.startswith("__")]

    header_data = ["序号", *(header_data if header_data else default_columns)]
    hidden_header = ["#", *(hidden_header if hidden_header else default_columns)]

    # self.export_column_width
    df_len_max = []
    for ele in header_data:
        w = get_string_len(ele)
        if w > max_column_width:
            w = max_column_width
        if w < default_column_width:
            w = default_column_width
        df_len_max.append(w)

    wb = Workbook()
    ws = wb.active
    row = get_column_letter(len(header_data))
    column = 1
    ws.append(header_data)
    for index, results in enumerate(data):
        results_list = []
        for h_index, h_item in enumerate(hidden_header):
            for key, val in results.items():
                if key == h_item:
                    if val is None or val == "":
                        results_list.append("")
                    else:
                        results_list.append(
                            val
                            if not (isinstance(val, dict) or isinstance(val, list))
                            else json.dumps(val, ensure_ascii=False)
                        )
                    # 计算最大列宽度
                    # print(f"@sting {key}={val}")
                    result_column_width = (
                        max_column_width
                        if (isinstance(val, dict) or isinstance(val, list))
                        else get_string_len(val)
                    )
                    if h_index != 0 and result_column_width > df_len_max[h_index]:
                        df_len_max[h_index] = (
                            result_column_width
                            if result_column_width < max_column_width
                            else max_column_width
                        )
        ws.append([index + 1, *results_list])
        column += 1
    # 　更新列宽
    for index, width in enumerate(df_len_max):
        ws.column_dimensions[get_column_letter(index + 1)].width = width
    tab = Table(displayName="Table", ref=f"A1:{row}{column}")  # 名称管理器
    style = TableStyleInfo(
        name="TableStyleLight11",
        showFirstColumn=True,
        showLastColumn=True,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    wb.save(fn)
    return True
